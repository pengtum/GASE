import openpyxl
wb = openpyxl.load_workbook(
    '../03_simulation_experiment/results/ModelTime.xlsx',
    data_only=True)
ws = wb['performance']
print(f'Sheet rows: {ws.max_row}\n')
hdr = f'{"#":>3}  {"Model":<42}  {"R-SSR":>8}  {"V-SSR":>8}  {"A-SSR":>8}  {"R-RandI":>7}  {"V-RandI":>7}  {"A-RandI":>7}'
print(hdr)
print('-' * len(hdr))
for r in range(3, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 18)]
    if row[1] is None:
        continue
    name = row[1] if isinstance(row[1], str) else str(row[1])
    if len(name) > 40:
        name = name[:38] + '..'
    def fmt(v, w=8, p=2):
        try:
            return f'{float(v):>{w}.{p}f}'
        except (TypeError, ValueError):
            return ' ' * w
    print(f'{(row[0] if row[0] is not None else ""):>3}  {name:<42}  '
          f'{fmt(row[2])}  {fmt(row[7])}  {fmt(row[12])}  '
          f'{fmt(row[3], 7, 3)}  {fmt(row[8], 7, 3)}  {fmt(row[13], 7, 3)}')
